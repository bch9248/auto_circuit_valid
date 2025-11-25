"""
Image retrieval methods for finding relevant schematic regions from PDF documents.

This module provides functionality to:
1. Extract text from extracted_text column in Excel
2. Find best matching windows in PDF schematics
3. Return top K cropped images
4. Save results to Excel
"""

import fitz  # PyMuPDF
import re
import os
import json
from collections import defaultdict
from typing import Dict, List, Tuple, Set, Optional, Any
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import io


class PDFWindowExtractor:
    """Extract best matching windows from PDF schematics based on target text."""
    
    def __init__(self, pdf_path: str, corpus_cache_path: Optional[str] = None):
        """
        Initialize PDF extractor.
        
        Args:
            pdf_path: Path to PDF file
            corpus_cache_path: Optional path to cached corpus JSON to avoid re-extraction
        """
        self.pdf_path = pdf_path
        self.pdf_document = fitz.open(pdf_path)
        self.corpus = {}
        self.corpus_cache_path = corpus_cache_path
        
        # Load or extract corpus
        if corpus_cache_path and os.path.exists(corpus_cache_path):
            self.load_corpus(corpus_cache_path)
            print(f"Loaded corpus from cache: {corpus_cache_path}")
        else:
            print("Extracting text from PDF...")
            self.extract_all_pages()
            if corpus_cache_path:
                self.save_corpus(corpus_cache_path)
    
    def extract_all_pages(self):
        """Extract text and positions from all pages of the PDF."""
        print(f"Extracting text from {len(self.pdf_document)} pages...")
        
        for page_num in range(len(self.pdf_document)):
            page = self.pdf_document[page_num]
            page_data = {
                'words': [],
                'formatted_text': []
            }
            
            # Extract words with positions
            words = page.get_text("words")
            for word in words:
                page_data['words'].append({
                    'text': word[4],
                    'bbox': word[:4],
                })
            
            # Extract formatted text with spans
            text_dict = page.get_text("dict")
            for block in text_dict["blocks"]:
                if "lines" in block:
                    for line in block["lines"]:
                        for span in line["spans"]:
                            page_data['formatted_text'].append({
                                'text': span['text'],
                                'bbox': span['bbox']
                            })
            
            self.corpus[page_num] = page_data
            
            if (page_num + 1) % 10 == 0:
                print(f"  Processed {page_num + 1} pages...")
        
        print(f"Extraction complete! Processed {len(self.pdf_document)} pages.")
        return self.corpus
    
    def normalize_text(self, text: str) -> str:
        """Normalize text for better matching."""
        text = re.sub(r'\s+', ' ', text.strip())
        text = re.sub(r'[^\w\s\.\#\-\_\%]', '', text)
        return text.upper()
    
    def calculate_page_confidence_with_coords(self, page_num: int, target_strings: List[str]) -> Tuple[float, List[str], List[Dict]]:
        """Calculate confidence score and return matched text coordinates."""
        if page_num not in self.corpus:
            return 0, [], []
        
        page_data = self.corpus[page_num]
        normalized_targets = [self.normalize_text(target) for target in target_strings]
        
        matched_strings = []
        matched_coords = []
        confidence_score = 0
        
        # Create lookup for all text elements with coordinates
        text_elements = []
        
        for word in page_data['words']:
            text_elements.append({
                'text': word['text'],
                'normalized': self.normalize_text(word['text']),
                'bbox': word['bbox'],
                'type': 'word'
            })
        
        for span in page_data['formatted_text']:
            text_elements.append({
                'text': span['text'],
                'normalized': self.normalize_text(span['text']),
                'bbox': span['bbox'],
                'type': 'span'
            })
        
        for i, target in enumerate(normalized_targets):
            matched = False
            
            # Exact match
            for element in text_elements:
                if target == element['normalized']:
                    matched_strings.append(target_strings[i])
                    matched_coords.append({
                        'original_text': target_strings[i],
                        'matched_text': element['text'],
                        'bbox': element['bbox'],
                        'match_type': 'exact',
                        'score': 1.0
                    })
                    confidence_score += 1
                    matched = True
                    break
            
            if matched:
                continue
            
            # Partial match
            for element in text_elements:
                if target in element['normalized'] or element['normalized'] in target:
                    matched_strings.append(target_strings[i])
                    matched_coords.append({
                        'original_text': target_strings[i],
                        'matched_text': element['text'],
                        'bbox': element['bbox'],
                        'match_type': 'partial',
                        'score': 0.5
                    })
                    confidence_score += 0.5
                    matched = True
                    break
            
            if matched:
                continue
            
            # Fuzzy match
            target_words = set(target.split())
            for element in text_elements:
                element_words = set(element['normalized'].split())
                if target_words.intersection(element_words):
                    overlap_ratio = len(target_words.intersection(element_words)) / len(target_words)
                    if overlap_ratio >= 0.5:
                        matched_strings.append(target_strings[i])
                        matched_coords.append({
                            'original_text': target_strings[i],
                            'matched_text': element['text'],
                            'bbox': element['bbox'],
                            'match_type': 'fuzzy',
                            'score': overlap_ratio * 0.3
                        })
                        confidence_score += overlap_ratio * 0.3
                        matched = True
                        break
        
        return confidence_score, matched_strings, matched_coords
    
    def sliding_window_analysis(self, page_num: int, target_strings: List[str], 
                               window_size=(1400, 500), overlap=0.3, dpi=300):
        """Analyze page using sliding window to find best regions."""
        
        page = self.pdf_document[page_num]
        zoom = dpi / 72
        
        # Generate page image
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False, colorspace=fitz.csRGB)
        img_width = pix.width
        img_height = pix.height
        
        # Get matched coordinates
        confidence, matched_strings, matched_coords = self.calculate_page_confidence_with_coords(page_num, target_strings)
        
        if not matched_coords:
            return []
        
        # Convert PDF coordinates to image coordinates
        image_coords = []
        for match in matched_coords:
            pdf_bbox = match['bbox']
            
            img_x0 = int(pdf_bbox[0] * zoom)
            img_y0 = int(pdf_bbox[1] * zoom)
            img_x1 = int(pdf_bbox[2] * zoom)
            img_y1 = int(pdf_bbox[3] * zoom)
            
            center_x = (img_x0 + img_x1) // 2
            center_y = (img_y0 + img_y1) // 2
            
            image_coords.append({
                'original_text': match['original_text'],
                'matched_text': match['matched_text'],
                'bbox': (img_x0, img_y0, img_x1, img_y1),
                'center': (center_x, center_y),
                'match_type': match['match_type'],
                'score': match['score']
            })
        
        # Generate sliding windows
        window_width, window_height = window_size
        step_x = int(window_width * (1 - overlap))
        step_y = int(window_height * (1 - overlap))
        
        windows = []
        window_id = 0
        
        x_positions = list(range(0, max(1, img_width - window_width + 1), step_x))
        if not x_positions or x_positions[-1] + window_width < img_width:
            x_positions.append(max(0, img_width - window_width))
        
        y_positions = list(range(0, max(1, img_height - window_height + 1), step_y))
        if not y_positions or y_positions[-1] + window_height < img_height:
            y_positions.append(max(0, img_height - window_height))
        
        for y in y_positions:
            for x in x_positions:
                actual_x = max(0, min(x, img_width - window_width))
                actual_y = max(0, min(y, img_height - window_height))
                
                window_bbox = (actual_x, actual_y, 
                             actual_x + window_width, 
                             actual_y + window_height)
                
                matches_in_window = []
                window_score = 0
                
                for coord in image_coords:
                    text_center = coord['center']
                    
                    if (window_bbox[0] <= text_center[0] <= window_bbox[2] and 
                        window_bbox[1] <= text_center[1] <= window_bbox[3]):
                        matches_in_window.append(coord)
                        window_score += coord['score']
                
                if matches_in_window:
                    window_info = {
                        'window_id': window_id,
                        'bbox': window_bbox,
                        'matches': matches_in_window,
                        'match_count': len(matches_in_window),
                        'window_score': window_score,
                        'coverage_ratio': len(matches_in_window) / len(target_strings)
                    }
                    windows.append(window_info)
                
                window_id += 1
        
        windows.sort(key=lambda w: (w['window_score'], w['match_count']), reverse=True)
        return windows
    
    def find_best_pages(self, target_strings: List[str], top_k: int = 2) -> List[Tuple[int, float]]:
        """Find top K pages with highest confidence for target strings."""
        page_scores = []
        
        for page_num in range(len(self.pdf_document)):
            confidence, matched_strings, _ = self.calculate_page_confidence_with_coords(page_num, target_strings)
            if confidence > 0:
                page_scores.append((page_num, confidence))
        
        page_scores.sort(key=lambda x: x[1], reverse=True)
        return page_scores[:top_k]
    
    def get_top_windows(self, target_strings: List[str], 
                       top_pages: int = 2, 
                       top_k: int = 5,
                       window_size=(1400, 500), 
                       overlap=0.3, 
                       dpi=300) -> List[Dict]:
        """
        Get top K windows across best pages for target text strings.
        
        Args:
            target_strings: List of text strings to search for
            top_pages: Number of top pages to analyze
            top_k: Final number of top windows to return
            window_size: Tuple of (width, height) for sliding window
            overlap: Overlap ratio between windows (0-1)
            dpi: Resolution for image extraction
            
        Returns:
            List of dictionaries containing window info and cropped images
        """
        print(f"\nSearching for {len(target_strings)} target strings in PDF...")
        
        # Find best pages
        best_pages = self.find_best_pages(target_strings, top_k=top_pages)
        
        if not best_pages:
            print("No matching pages found!")
            return []
        
        print(f"Found {len(best_pages)} pages with matches")
        
        # Collect windows from all best pages
        all_windows = []
        
        for page_num, page_score in best_pages:
            print(f"  Analyzing page {page_num} (score: {page_score:.2f})...")
            
            windows = self.sliding_window_analysis(
                page_num=page_num,
                target_strings=target_strings,
                window_size=window_size,
                overlap=overlap,
                dpi=dpi
            )
            
            # Add page info and generate cropped images
            for window in windows:
                window['page_num'] = page_num
                window['page_score'] = page_score
                window['combined_score'] = window['window_score'] + (page_score * 0.1)
                
                # Generate cropped image
                cropped_img = self._crop_window_from_page(page_num, window['bbox'], dpi)
                window['cropped_image'] = cropped_img
                
                all_windows.append(window)
        
        # Sort by combined score and take top K
        all_windows.sort(key=lambda w: (w['combined_score'], w['window_score']), reverse=True)
        top_windows = all_windows[:top_k]
        
        print(f"\nReturning top {len(top_windows)} windows:")
        for i, window in enumerate(top_windows, 1):
            print(f"  {i}. Page {window['page_num']}, Score: {window['window_score']:.2f}, "
                  f"Matches: {window['match_count']}")
        
        return top_windows
    
    def _crop_window_from_page(self, page_num: int, bbox: Tuple[int, int, int, int], dpi: int = 300) -> Image.Image:
        """Crop a window region from a PDF page."""
        page = self.pdf_document[page_num]
        zoom = dpi / 72
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False, colorspace=fitz.csRGB)
        
        # Convert pixmap to PIL Image
        img_data = pix.tobytes("png")
        full_img = Image.open(io.BytesIO(img_data))
        
        # Crop window
        cropped_img = full_img.crop(bbox)
        return cropped_img
    
    def save_corpus(self, filename: str):
        """Save the corpus to a JSON file."""
        with open(filename, 'w') as f:
            json.dump(self.corpus, f, indent=2)
        print(f"Corpus saved to {filename}")
    
    def load_corpus(self, filename: str):
        """Load corpus from a JSON file."""
        with open(filename, 'r') as f:
            self.corpus = json.load(f)
    
    def close(self):
        """Close the PDF document."""
        self.pdf_document.close()


def extract_target_strings_from_excel(excel_path: str, 
                                      sheet_name: str = None,
                                      extracted_text_column: str = 'I',
                                      start_row: int = 2) -> Dict[int, List[str]]:
    """
    Extract target text strings from Excel file's Extracted_text column.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Sheet name (uses active sheet if None)
        extracted_text_column: Column containing extracted text JSON arrays
        start_row: Row to start reading from (1-indexed)
        
    Returns:
        Dictionary mapping row numbers to lists of target strings
    """
    print(f"Loading target strings from: {excel_path}")
    
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name] if sheet_name else wb.active
    
    target_texts_dict = {}
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
        # Get the extracted text cell
        col_idx = ord(extracted_text_column.upper()) - ord('A')
        cell_value = row[col_idx].value
        
        if cell_value:
            try:
                # Parse JSON array
                text_list = json.loads(cell_value)
                if text_list:
                    target_texts_dict[row_idx] = text_list
            except json.JSONDecodeError:
                print(f"Warning: Could not parse JSON in row {row_idx}")
                continue
    
    wb.close()
    print(f"Loaded {len(target_texts_dict)} target text sets")
    return target_texts_dict


def save_windows_to_excel(windows_dict: Dict[int, List[Dict]], 
                          output_path: str,
                          source_excel_path: str = None,
                          sheet_name: str = None):
    """
    Save cropped windows to Excel file with images.
    
    Args:
        windows_dict: Dictionary mapping row numbers to lists of window dicts
        output_path: Output Excel file path
        source_excel_path: Optional source Excel to copy data from
        sheet_name: Sheet name for source Excel
    """
    print(f"\nSaving results to: {output_path}")
    
    # Load source workbook if provided, otherwise create new
    if source_excel_path and os.path.exists(source_excel_path):
        wb = load_workbook(source_excel_path)
        sheet = wb[sheet_name] if sheet_name else wb.active
        # Add new columns for window images
        start_col = sheet.max_column + 1
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Results"
        start_col = 1
        
        # Create headers
        headers = ['Row', 'Page', 'Window_ID', 'Score', 'Match_Count', 'Matched_Texts', 
                   'Window_1', 'Window_2', 'Window_3', 'Window_4', 'Window_5']
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx, value=header)
        start_col = 7  # Start images at column 7
    
    output_row = 2 if not source_excel_path else 2
    
    for row_num, windows in sorted(windows_dict.items()):
        for window_idx, window in enumerate(windows, 1):
            if not source_excel_path:
                # Write window metadata
                sheet.cell(row=output_row, column=1, value=row_num)
                sheet.cell(row=output_row, column=2, value=window['page_num'])
                sheet.cell(row=output_row, column=3, value=window['window_id'])
                sheet.cell(row=output_row, column=4, value=round(window['window_score'], 2))
                sheet.cell(row=output_row, column=5, value=window['match_count'])
                
                matched_texts = ', '.join([m['original_text'] for m in window['matches']])
                sheet.cell(row=output_row, column=6, value=matched_texts)
            
            # Save cropped image temporarily
            img_buffer = io.BytesIO()
            window['cropped_image'].save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Insert image into Excel
            excel_img = ExcelImage(img_buffer)
            excel_img.width = 300
            excel_img.height = int(300 * window['cropped_image'].height / window['cropped_image'].width)
            
            cell_ref = sheet.cell(row=output_row, column=start_col + window_idx - 1).coordinate
            sheet.add_image(excel_img, cell_ref)
            
            # Adjust row height
            sheet.row_dimensions[output_row].height = excel_img.height * 0.75
        
        output_row += 1
    
    wb.save(output_path)
    print(f"✓ Saved results with {len(windows_dict)} rows to: {output_path}")


def find_image_crop(target_strings: List[str],
                   pdf_path: str,
                   top_k: int = 5,
                   window_size: Tuple[int, int] = (1400, 500),
                   corpus_cache_path: Optional[str] = None,
                   verbose: bool = True) -> Dict[str, Any]:
    """
    Find and crop top K matching windows from PDF based on target strings.
    This function is designed to be called from query_executor.
    
    Args:
        target_strings: List of text strings to search for in PDF
        pdf_path: Path to PDF schematic
        top_k: Number of top windows to return
        window_size: (width, height) for sliding window
        corpus_cache_path: Optional path to cache PDF text extraction
        verbose: Print detailed output
        
    Returns:
        Dictionary containing:
        - success: bool
        - windows: List of window dicts with cropped images
        - num_windows: int
        - target_strings_count: int
    """
    if verbose:
        print(f"\nSearching PDF for {len(target_strings)} target strings...")
    
    try:
        # Initialize PDF extractor
        extractor = PDFWindowExtractor(pdf_path, corpus_cache_path)
        
        # Get top windows
        top_windows = extractor.get_top_windows(
            target_strings=target_strings,
            top_pages=2,
            top_k=top_k,
            window_size=window_size,
            overlap=0.3,
            dpi=300
        )
        
        # Cleanup
        extractor.close()
        
        if verbose:
            if top_windows:
                print(f"✓ Found {len(top_windows)} matching windows")
            else:
                print("⚠ No matching windows found")
        
        return {
            'success': True,
            'windows': top_windows,
            'num_windows': len(top_windows),
            'target_strings_count': len(target_strings),
            'pdf_path': pdf_path
        }
        
    except Exception as e:
        if verbose:
            print(f"✗ Error in image retrieval: {str(e)}")
        
        return {
            'success': False,
            'error': str(e),
            'windows': [],
            'num_windows': 0,
            'target_strings_count': len(target_strings)
        }


def process_excel_with_pdf_retrieval(excel_path: str,
                                     pdf_path: str,
                                     output_path: str,
                                     sheet_name: str = None,
                                     extracted_text_column: str = 'I',
                                     start_row: int = 2,
                                     top_k: int = 5,
                                     window_size: Tuple[int, int] = (1400, 500),
                                     corpus_cache_path: Optional[str] = None):
    """
    Main pipeline function: Extract target texts from Excel, find matching PDF windows, save results.
    
    Args:
        excel_path: Path to input Excel with extracted text
        pdf_path: Path to PDF schematic
        output_path: Path for output Excel
        sheet_name: Sheet name in Excel (None = active sheet)
        extracted_text_column: Column containing target text JSON arrays
        start_row: First data row (1-indexed)
        top_k: Number of top windows to extract per row
        window_size: (width, height) for sliding window
        corpus_cache_path: Optional path to cache PDF text extraction
    """
    print("="*60)
    print("PDF SCHEMATIC WINDOW RETRIEVAL PIPELINE")
    print("="*60)
    
    # Step 1: Extract target strings from Excel
    target_texts_dict = extract_target_strings_from_excel(
        excel_path, sheet_name, extracted_text_column, start_row
    )
    
    if not target_texts_dict:
        print("No target texts found in Excel!")
        return
    
    # Step 2: Initialize PDF extractor
    extractor = PDFWindowExtractor(pdf_path, corpus_cache_path)
    
    # Step 3: Process each row's target texts
    windows_results = {}
    
    for row_num, target_strings in target_texts_dict.items():
        print(f"\n{'='*60}")
        print(f"Processing Row {row_num} with {len(target_strings)} target strings")
        print(f"{'='*60}")
        
        top_windows = extractor.get_top_windows(
            target_strings=target_strings,
            top_pages=2,
            top_k=top_k,
            window_size=window_size,
            overlap=0.3,
            dpi=300
        )
        
        if top_windows:
            windows_results[row_num] = top_windows
    
    # Step 4: Save results to Excel
    if windows_results:
        save_windows_to_excel(
            windows_dict=windows_results,
            output_path=output_path,
            source_excel_path=excel_path,
            sheet_name=sheet_name
        )
    
    # Cleanup
    extractor.close()
    
    print("\n" + "="*60)
    print("PIPELINE COMPLETE")
    print("="*60)


# Example usage
if __name__ == "__main__":
    # Example: Process extracted queries Excel
    process_excel_with_pdf_retrieval(
        excel_path='output_extracted_data.xlsx',
        pdf_path='Input/G12_MACHU14_TLD_1217/G12_MACHU14_TLD_1217.pdf',
        output_path='output_with_pdf_windows.xlsx',
        sheet_name=None,  # Use active sheet
        extracted_text_column='I',
        start_row=2,
        top_k=5,
        window_size=(1400, 500),
        corpus_cache_path='pdf_corpus_cache.json'
    )
