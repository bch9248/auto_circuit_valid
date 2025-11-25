from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import os
import shutil
from typing import Dict, List, Optional
from PIL import Image as PILImage
import io
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from typing import Dict, List, Optional, Tuple

class ExcelExtractor:
    """Extract data and images from Excel files."""
    
    def __init__(self, input_file: str, sheet_name: Optional[str] = None):
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        
    def load_workbook_data(self) -> None:
        """Load the workbook and select the appropriate sheet."""
        print(f"Loading file: {self.input_file}")
        self.workbook = load_workbook(self.input_file)
        print(f"\nAvailable sheets: {self.workbook.sheetnames}")
        
        if self.sheet_name and self.sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[self.sheet_name]
        else:
            self.sheet = self.workbook.active
            
        print(f"\nUsing sheet: {self.sheet.title}")
    
    def extract_text_data(self, start_row: int, end_row: int, columns: List[str]) -> List[List]:
        """Extract text data from specified range."""
        data_rows = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in columns:
                cell_value = self.sheet[f'{col}{row}'].value
                row_data.append(cell_value if cell_value is not None else '')
            data_rows.append(row_data)
        
        print(f"Extracted {len(data_rows)} rows of text data")
        return data_rows
    
    def extract_images(self, start_row: int, end_row: int, target_column: int, visualize: bool = False) -> Dict[int, Tuple[bytes, str]]:
        """Extract images and return as byte data with format to avoid file closure issues."""
        print(f"\n{'='*60}")
        print(f"DEBUG: Extracting images")
        print(f"{'='*60}")
        print(f"Target range: rows {start_row}-{end_row}")
        print(f"Target column: {target_column} ({self._get_column_letter(target_column)})")
        
        # Step 1: Extract all images with their data cached
        all_images = {}
        image_preview_data = []
        
        print(f"\nTotal images in sheet: {len(self.sheet._images)}")
        
        for idx, img in enumerate(self.sheet._images, 1):
            if not hasattr(img, 'anchor'):
                continue
            
            anchor = img.anchor
            
            if hasattr(anchor, '_from'):
                from_cell = anchor._from
                row_num = from_cell.row
                col_num = from_cell.col + 1  # Convert to 1-based
                
                try:
                    # Cache the image data immediately to avoid file closure issues
                    image_data = img._data()
                    img_format = getattr(img, 'format', 'png').lower()
                    
                    # Store as tuple: (byte_data, format)
                    all_images[row_num] = (image_data, img_format)
                    
                    # Get image dimensions for logging
                    pil_img = PILImage.open(io.BytesIO(image_data))
                    width, height = pil_img.size
                    
                    # print(f"  Image #{idx} at Row {row_num}, Col {self._get_column_letter(col_num)}: {width}x{height} ({img_format})")
                    
                    image_preview_data.append({
                        'index': idx,
                        'row': row_num,
                        'col': col_num,
                        'data': image_data,
                        'format': img_format
                    })
                    
                except Exception as e:
                    print(f"  ⚠ ERROR extracting image #{idx} at row {row_num}: {str(e)}")
                    continue
        
        print(f"\nTotal images extracted: {len(all_images)}")
        print(f"Rows with images: {sorted(all_images.keys())}")
        
        # Step 2: Filter images by target range and column
        print(f"\n{'='*60}")
        print(f"Filtering images...")
        print(f"{'='*60}")
        
        relevant_images = {}
        matched_preview_data = []
        
        for row_num in range(start_row, end_row + 1):
            if row_num in all_images:
                # Check if we need to validate column (we can't do this after caching)
                # So we'll include all images in the row range
                relevant_images[row_num] = all_images[row_num]
                print(f"  ✓ Row {row_num} - MATCHED")
                
                for item in image_preview_data:
                    if item['row'] == row_num:
                        matched_preview_data.append(item)
                        break
        
        print(f"\nSummary:")
        print(f"  - Total images extracted: {len(all_images)}")
        print(f"  - Images in target range: {len(relevant_images)}")
        print(f"  - Matched rows: {sorted(relevant_images.keys())}")
        
        missing_rows = [r for r in range(start_row, end_row + 1) if r not in relevant_images]
        if missing_rows:
            print(f"  - Rows without images: {missing_rows[:10]}{'...' if len(missing_rows) > 10 else ''}")
        print(f"{'='*60}\n")
        
        if visualize and matched_preview_data:
            self._visualize_images(matched_preview_data, start_row, end_row, target_column)
        
        return relevant_images
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to Excel column letter."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + 65) + result
            col_num //= 26
        return result
    
    def _visualize_images(self, image_data: List[Dict], start_row: int, end_row: int, target_column: int) -> None:
        """Visualize detected images in a grid."""
        print(f"\nGenerating image preview...")
        
        num_images = len(image_data)
        if num_images == 0:
            print("No images to visualize")
            return
        
        # Calculate grid size
        cols = min(4, num_images)
        rows = (num_images + cols - 1) // cols
        
        fig, axes = plt.subplots(rows, cols, figsize=(15, 4 * rows))
        if num_images == 1:
            axes = [axes]
        else:
            axes = axes.flatten() if rows > 1 else axes
        
        for idx, (ax, img_info) in enumerate(zip(axes[:num_images], image_data)):
            try:
                # Extract image data
                img = img_info['img']
                image_data_bytes = img._data()
                
                # Convert to PIL Image
                pil_img = PILImage.open(io.BytesIO(image_data_bytes))
                
                # Display image
                ax.imshow(pil_img)
                ax.axis('off')
                
                # Add title with position info
                title = f"Image #{img_info['index']}\n"
                title += f"Row {img_info['row']}, Col {self._get_column_letter(img_info['col'])}"
                
                # Highlight if in target range
                in_range = start_row <= img_info['row'] <= end_row
                color = 'green' if in_range else 'red'
                ax.set_title(title, fontsize=10, color=color, weight='bold')
                
                # Add border
                for spine in ax.spines.values():
                    spine.set_edgecolor(color)
                    spine.set_linewidth(3)
                
            except Exception as e:
                ax.text(0.5, 0.5, f'Error loading image\n{str(e)}', 
                       ha='center', va='center', transform=ax.transAxes)
                ax.axis('off')
        
        # Hide unused subplots
        for ax in axes[num_images:]:
            ax.axis('off')
        
        plt.suptitle(f'Detected Images (Target: Column {self._get_column_letter(target_column)}, Rows {start_row}-{end_row})',
                    fontsize=14, weight='bold')
        plt.tight_layout()
        
        # Save figure
        preview_file = 'image_detection_preview.png'
        plt.savefig(preview_file, dpi=150, bbox_inches='tight')
        print(f"✓ Image preview saved to: {preview_file}")
        
        # Show figure
        plt.show()
    def examine_image_locations(self, visualize: bool = True) -> List[Dict]:
        """
        Examine all images in the sheet and return their locations.
        
        Args:
            visualize: Whether to display a visual grid showing image locations
            
        Returns:
            List of dictionaries containing image location information
        """
        print(f"\n{'='*60}")
        print(f"EXAMINING IMAGE LOCATIONS")
        print(f"{'='*60}")
        print(f"Sheet: {self.sheet.title}")
        
        image_locations = []
        
        # Get sheet dimensions
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
        print(f"Total images in sheet: {len(self.sheet._images)}\n")
        
        # Extract location info for each image
        for idx, img in enumerate(self.sheet._images, 1):
            if not hasattr(img, 'anchor'):
                continue
                
            anchor = img.anchor
            
            if hasattr(anchor, '_from'):
                from_cell = anchor._from
                row_num = from_cell.row
                col_num = from_cell.col + 1  # Convert to 1-based
                col_letter = self._get_column_letter(col_num)
                
                try:
                    # Get image data and properties
                    image_data = img._data()
                    img_format = getattr(img, 'format', 'png').lower()
                    pil_img = PILImage.open(io.BytesIO(image_data))
                    width, height = pil_img.size
                    
                    location_info = {
                        'index': idx,
                        'row': row_num,
                        'col': col_num,
                        'col_letter': col_letter,
                        'cell': f'{col_letter}{row_num}',
                        'width': width,
                        'height': height,
                        'format': img_format,
                        'image_data': image_data
                    }
                    
                    image_locations.append(location_info)
                    
                    print(f"Image #{idx:2d} | Cell: {location_info['cell']:6s} | "
                          f"Row: {row_num:3d} | Col: {col_letter:3s} ({col_num:2d}) | "
                          f"Size: {width:4d}x{height:4d} | Format: {img_format}")
                    
                except Exception as e:
                    print(f"  ⚠ ERROR examining image #{idx}: {str(e)}")
                    continue
        
        print(f"\n{'='*60}")
        print(f"Total images examined: {len(image_locations)}")
        print(f"{'='*60}\n")
        
        if visualize and image_locations:
            self._visualize_image_grid(image_locations, max_row, max_col)
        
        return image_locations
    
    def _visualize_image_grid(self, image_locations: List[Dict], max_row: int, max_col: int) -> None:
        """
        Visualize image locations on a grid representation of the Excel sheet.
        """
        print("Generating image location visualization...")
        
        fig = plt.figure(figsize=(16, 10))
        
        # Create two subplots: grid view and image preview
        gs = fig.add_gridspec(2, 1, height_ratios=[1, 2], hspace=0.3)
        
        # Subplot 1: Grid showing image positions
        ax_grid = fig.add_subplot(gs[0])
        
        # Limit displayed range for better visibility
        display_rows = min(max_row, 100)
        display_cols = min(max_col, 26)
        
        # Create grid
        grid = [[0 for _ in range(display_cols)] for _ in range(display_rows)]
        
        # Mark cells with images
        for img_info in image_locations:
            if img_info['row'] <= display_rows and img_info['col'] <= display_cols:
                grid[img_info['row']-1][img_info['col']-1] = img_info['index']
        
        # Plot grid
        ax_grid.imshow([[1 if cell == 0 else 0 for cell in row] for row in grid], 
                       cmap='Greys', alpha=0.3, aspect='auto')
        
        # Add image markers and labels
        for img_info in image_locations:
            if img_info['row'] <= display_rows and img_info['col'] <= display_cols:
                x = img_info['col'] - 1
                y = img_info['row'] - 1
                
                # Draw circle at image location
                circle = plt.Circle((x, y), 0.4, color='red', fill=True, alpha=0.7)
                ax_grid.add_patch(circle)
                
                # Add label
                ax_grid.text(x, y, str(img_info['index']), 
                            ha='center', va='center', color='white', 
                            fontsize=8, fontweight='bold')
        
        # Configure grid axes
        ax_grid.set_xlim(-0.5, display_cols - 0.5)
        ax_grid.set_ylim(display_rows - 0.5, -0.5)
        
        # Set ticks
        ax_grid.set_xticks(range(display_cols))
        ax_grid.set_xticklabels([self._get_column_letter(i+1) for i in range(display_cols)])
        ax_grid.set_yticks(range(0, display_rows, max(1, display_rows//20)))
        ax_grid.set_yticklabels(range(1, display_rows+1, max(1, display_rows//20)))
        
        ax_grid.set_xlabel('Column', fontsize=10, fontweight='bold')
        ax_grid.set_ylabel('Row', fontsize=10, fontweight='bold')
        ax_grid.set_title(f'Image Locations in Sheet: {self.sheet.title}', 
                         fontsize=12, fontweight='bold', pad=10)
        ax_grid.grid(True, alpha=0.2)
        
        # Subplot 2: Image thumbnails
        ax_images = fig.add_subplot(gs[1])
        ax_images.axis('off')
        
        # Display image thumbnails
        num_images = len(image_locations)
        cols = min(6, num_images)
        rows = (num_images + cols - 1) // cols
        
        inner_gs = gs[1].subgridspec(rows, cols, wspace=0.2, hspace=0.3)
        
        for idx, img_info in enumerate(image_locations):
            row_idx = idx // cols
            col_idx = idx % cols
            
            ax = fig.add_subplot(inner_gs[row_idx, col_idx])
            
            try:
                pil_img = PILImage.open(io.BytesIO(img_info['image_data']))
                ax.imshow(pil_img)
                ax.axis('off')
                
                title = f"#{img_info['index']}: {img_info['cell']}\n{img_info['width']}x{img_info['height']}"
                ax.set_title(title, fontsize=8, fontweight='bold')
                
                # Add colored border
                for spine in ax.spines.values():
                    spine.set_edgecolor('red')
                    spine.set_linewidth(2)
                    
            except Exception as e:
                ax.text(0.5, 0.5, f'Error\n{img_info["cell"]}', 
                       ha='center', va='center', fontsize=8)
                ax.axis('off')
        
        plt.suptitle('Image Location Analysis', fontsize=14, fontweight='bold', y=0.98)
        
        # Save figure
        output_file = 'image_location_analysis.png'
        plt.savefig(output_file, dpi=150, bbox_inches='tight')
        print(f"✓ Image location analysis saved to: {output_file}")
        
        plt.show()
    def extract_all_column_images(self, column: int, output_dir: str = 'all_column_images') -> None:
        """Extract and save ALL images from a specific column."""
        print(f"\n{'='*60}")
        print(f"EXTRACTING ALL IMAGES FROM COLUMN {self._get_column_letter(column)} ({column})")
        print(f"{'='*60}")
        
        # Create output directory
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir, exist_ok=True)
        print(f"Created directory: {output_dir}")
        
        # Extract all images from the column
        column_images = []
        
        for idx, img in enumerate(self.sheet._images, 1):
            if not hasattr(img, 'anchor'):
                continue
            
            anchor = img.anchor
            
            if hasattr(anchor, '_from'):
                from_cell = anchor._from
                row_num = from_cell.row
                col_num = from_cell.col + 1  # Convert to 1-based
                
                # Check if this image is in the target column
                if col_num == column:
                    try:
                        # Get image data
                        image_data = img._data()
                        img_format = getattr(img, 'format', 'png').lower()
                        
                        # Get dimensions
                        pil_img = PILImage.open(io.BytesIO(image_data))
                        width, height = pil_img.size
                        
                        # Validate format
                        ext = img_format.lower()
                        if ext not in ['png', 'jpg', 'jpeg', 'gif', 'bmp']:
                            ext = 'png'
                        
                        # Save image
                        filename = f'col_{self._get_column_letter(column)}_row_{row_num}.{ext}'
                        filepath = os.path.join(output_dir, filename)
                        
                        # Save based on format
                        if ext in ['jpg', 'jpeg']:
                            pil_img = pil_img.convert('RGB')
                            pil_img.save(filepath, 'JPEG', quality=95)
                        else:
                            pil_img.save(filepath, ext.upper())
                        
                        column_images.append({
                            'row': row_num,
                            'filepath': filepath,
                            'size': (width, height),
                            'format': img_format
                        })
                        
                        print(f"  ✓ Row {row_num}: Saved {filename} ({width}x{height})")
                        
                    except Exception as e:
                        print(f"  ✗ ERROR saving image at row {row_num}: {str(e)}")
                        continue
        
        print(f"\n{'='*60}")
        print(f"SUMMARY:")
        print(f"  - Total images saved: {len(column_images)}")
        print(f"  - Output directory: {output_dir}")
        print(f"  - Rows with images: {sorted([img['row'] for img in column_images])}")
        print(f"{'='*60}\n")
        
        return column_images
class ImageSaver:
    """Save images to directory."""
    
    def __init__(self, output_dir: str = 'output_images', max_dimension: int = 5000):
        self.output_dir = output_dir
        self.max_dimension = max_dimension
        
    def setup_directory(self) -> None:
        """Create output directory for images."""
        if os.path.exists(self.output_dir):
            shutil.rmtree(self.output_dir)
        os.makedirs(self.output_dir, exist_ok=True)
        print(f"\nCreated image output directory: {self.output_dir}")
    
    def save_image(self, image_data: bytes, img_format: str, row_index: int) -> str:
        """Save image from byte data to file with size validation and resizing."""
        try:
            pil_img = PILImage.open(io.BytesIO(image_data))
            
            width, height = pil_img.size
            print(f"  Processing image for row {row_index}: {width}x{height}")
            
            # Resize if too large
            if width > self.max_dimension or height > self.max_dimension:
                print(f"    Resizing large image...")
                
                if width > height:
                    new_width = self.max_dimension
                    new_height = int(height * (self.max_dimension / width))
                else:
                    new_height = self.max_dimension
                    new_width = int(width * (self.max_dimension / height))
                
                pil_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                print(f"    Resized to: {new_width}x{new_height}")
            
            # Validate format
            ext = img_format.lower()
            if ext not in ['png', 'jpg', 'jpeg', 'gif', 'bmp']:
                ext = 'png'
            
            filename = f'image_row_{row_index}.{ext}'
            filepath = os.path.join(self.output_dir, filename)
            
            # Save based on format
            if ext in ['jpg', 'jpeg']:
                pil_img = pil_img.convert('RGB')
                pil_img.save(filepath, 'JPEG', quality=95)
            else:
                pil_img.save(filepath, ext.upper())
            
            print(f"  ✓ Saved image: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"  ✗ ERROR saving image for row {row_index}: {str(e)}")
            return f"ERROR: {str(e)}"


class ExcelWriter:
    """Write extracted data to new Excel file."""
    
    def __init__(self, output_file: str):
        self.output_file = output_file
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        
    def write_data(self, data_rows: List[List], images_dict: Dict[int, Image], 
                   source_start_row: int, script_language_col_index: int,
                   image_saver: ImageSaver) -> int:
        """Write data and image paths to output sheet."""
        output_row = 1  # Start from row 1 (no header)
        
        for source_idx, row_data in enumerate(data_rows):
            source_row = source_idx + source_start_row
            
            # Check if Script_Language column has value (filter empty rows)
            script_language_value = row_data[script_language_col_index] if script_language_col_index < len(row_data) else ''
            if not script_language_value or str(script_language_value).strip() == '':
                continue
            
            # Column mapping:
            # Output A: Index (output_row)
            # Output B: Method (query{index})
            # Output C: Item (source B, index 0)
            # Output D: Description of Criteria (source C, index 1)
            # Output E: Pass/Fail Criteria (merge source D-I, indices 2-7)
            # Output F: Check Results (source J, index 8)
            # Output G: Script_Language (source M, index 10)
            # Output H: Image Path
            
            # A: Index
            if output_row==1:
                self.sheet.cell(row=output_row, column=1, value="Idx")
            else:
                self.sheet.cell(row=output_row, column=1, value=output_row)
            
            # B: Method
            if output_row==1:
                self.sheet.cell(row=output_row, column=2, value="Method")
            else:
                self.sheet.cell(row=output_row, column=2, value=f"query{output_row-1:04d}")
            # C: Item
            self.sheet.cell(row=output_row, column=3, value=row_data[0])
            
            # D: Description of Criteria
            self.sheet.cell(row=output_row, column=4, value=row_data[1])
            
            # E: Pass/Fail Criteria (merge D-I)
            pass_fail_values = []
            for pf_idx in range(4, 8):  # Indices 2-7 correspond to D-I
                if pf_idx < len(row_data) and row_data[pf_idx]:
                    pass_fail_values.append(str(row_data[pf_idx]))
            merged_value = ' '.join(pass_fail_values) if pass_fail_values else ''
            self.sheet.cell(row=output_row, column=5, value=merged_value)
            

            # F: Check Results  # check result is empty?????
            self.sheet.cell(row=output_row, column=6, value=row_data[9])
            
            # G: Script_Language
            self.sheet.cell(row=output_row, column=7, value=row_data[10])

            # H: Image Path
            if output_row==1:
                self.sheet.cell(row=output_row, column=8, value="Image_path")
            else:
                if source_row in images_dict:
                    print(f"at row {output_row} saves image number: {source_row}")
                    image_data, img_format = images_dict[source_row]
                    image_path = image_saver.save_image(image_data, img_format, output_row)
                    
                    self.sheet.cell(row=output_row, column=8, value=image_path)
                
            output_row += 1
        
        return output_row - 1  # Return actual number of rows written
    
    def format_sheet(self, num_rows: int) -> None:
        """Format column widths and row heights."""
        # Adjust column widths
        column_widths = {
            'A': 10,   # Index
            'B': 15,   # Method
            'C': 30,   # Item
            'D': 50,   # Description of Criteria
            'E': 50,   # Pass/Fail Criteria
            'F': 30,   # Check Results
            'G': 20,   # Script_Language
            'H': 40    # Image Path
        }
        for col_letter, width in column_widths.items():
            self.sheet.column_dimensions[col_letter].width = width
        
        # Set row heights
        for row_num in range(1, num_rows + 1):
            self.sheet.row_dimensions[row_num].height = 30
    
    def save(self) -> None:
        """Save the workbook."""
        self.workbook.save(self.output_file)
        print(f"\n✓ Successfully created {self.output_file}")
def main():
    """Main execution function."""
    # Configuration
    input_file = r'Input/G12_MACHU14_TLD_1217/SVTP803_Machu1416 CMIT_PV_CM01_Vendor.xlsm_Reply by Henry_2024_10_29_V2_20251031.xlsm'
    # input_file = r'Input/SVTP804 2nd Release Candidate_Cashmere_AMD_DB_CM01_250602/SVTP804 2nd Release Candidate_Cashmere_AMD_DB_CM01_250602.xlsm'
    sheet_name = 'WWAN '
    output_file = 'query_w_answer.xlsx'
    image_output_dir = 'output_images'
    
    # Data extraction parameters
    start_row = 18
    end_row = 64
    # B: No, C: Item, D-I: Pass/Fail, J: Category, K: Check Results, M: Script_Language
    columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M']
    image_column = 12  # Column L (1-based index)
    script_language_index = 10  # Index of 'M' column in the columns list (0-based)
    
    # Extract data
    print("=" * 60)
    print("STEP 1: Extracting data from source file")
    print("=" * 60)
    extractor = ExcelExtractor(input_file, sheet_name)
    extractor.load_workbook_data()
    
    # After loading the workbook, add this:
    print("\n" + "=" * 60)
    print("EXAMINING IMAGE LOCATIONS")
    print("=" * 60)
    # image_locations = extractor.examine_image_locations(visualize=True)
    # all_col_l_images = extractor.extract_all_column_images(image_column, 'all_column_L_images')
    
    # Optional: Print summary table
    # print("\nSummary Table:")
    # print(f"{'Index':<8} {'Cell':<10} {'Row':<6} {'Col':<8} {'Size':<15} {'Format':<10}")
    # print("-" * 70)
    # for img in image_locations:
    #     print(f"{img['index']:<8} {img['cell']:<10} {img['row']:<6} "
    #           f"{img['col_letter']:<8} {img['width']}x{img['height']:<10} {img['format']:<10}")
    
    data_rows = extractor.extract_text_data(start_row, end_row, columns)
    images = extractor.extract_images(start_row, end_row, image_column, visualize=False)
    
    # Setup image saver
    print("\n" + "=" * 60)
    print("STEP 2: Setting up image output directory")
    print("=" * 60)
    image_saver = ImageSaver(image_output_dir)
    image_saver.setup_directory()
    
    # Write output
    print("\n" + "=" * 60)
    print("STEP 3: Writing output file")
    print("=" * 60)
    image_start_row=17
    writer = ExcelWriter(output_file)
    num_rows = writer.write_data(data_rows, images, image_start_row, script_language_index, image_saver)
    
    writer.format_sheet(num_rows)
    writer.save()
    
    print("\n" + "=" * 60)
    print(f"✓ Process completed!")
    print(f"  - Output file: {output_file}")
    print(f"  - Images saved to: {image_output_dir}")
    print(f"  - Total rows processed: {num_rows}")
    print("=" * 60)
if __name__ == "__main__":
    main()